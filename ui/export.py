from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from database import Database

HEADER_FILL = PatternFill("solid", fgColor="1565C0")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SUMMARY_FILL = PatternFill("solid", fgColor="E8F5E9")
SUMMARY_FONT = Font(bold=True, color="1B5E20")
HEADERS = ["Date", "Day", "Project Name", "Code", "Description", "Billability", "Hours"]


def export_week_to_excel(db: Database, iso_year: int, iso_week: int, file_path: str):
    entries = db.get_entries_for_week(iso_year, iso_week)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Week {iso_year}-W{iso_week:02d}"

    # Headers
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    total_hours = 0.0
    for e in entries:
        d = date.fromisoformat(e["date"])
        ws.append([
            e["date"],
            d.strftime("%A"),
            e["project_name"],
            e["project_code"],
            e["description"],
            e["billability"],
            e["hours"],
        ])
        total_hours += e["hours"]

    # Summary row
    summary_row = ws.max_row + 1
    ws.cell(summary_row, 1, "TOTAL")
    ws.cell(summary_row, 7, total_hours)
    for col in range(1, 8):
        cell = ws.cell(summary_row, col)
        cell.font = SUMMARY_FONT
        cell.fill = SUMMARY_FILL

    # Column widths
    col_widths = [12, 12, 28, 12, 36, 14, 8]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = width

    wb.save(file_path)
